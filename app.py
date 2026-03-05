from flask import Flask, request, jsonify, render_template, send_file, after_this_request
import pandas as pd
import tempfile
import openpyxl
import time
import datetime
from pathlib import Path
from openpyxl import load_workbook
import os
import uuid

app = Flask(__name__)

TEMPLATE_PATH = Path("Entradas/TEMPLATE-SITUAÇÃO DE PROJETOS COMPLEMENTARES.xlsx")
UPLOAD_FOLDER = Path("temp_uploads")
UPLOAD_FOLDER.mkdir(exist_ok=True)
SAIDAS_FOLDER = Path("Saidas")
SAIDAS_FOLDER.mkdir(exist_ok=True)

def ajusta_lista(arquivo_excel, fases_selecionadas, disciplinas_selecionadas=None, data_range=None):
    if data_range is None:
        data_range=[datetime.date.today(), datetime.date.today().replace(year=datetime.date.today().year + 1)]
        
    tabela = pd.read_excel(arquivo_excel, skiprows=2)

    colunas_desejadas = ['Disciplina', 'Fase', 'Código', 'Revisão', 'Documento', 'Extensão', 'Situação', 'Título', 'Data de Alteração']
    tabela = tabela[tabela.columns.intersection(colunas_desejadas)].copy()

    tabela = tabela[tabela['Fase'].isin(fases_selecionadas)]

    if disciplinas_selecionadas:
        tabela = tabela[tabela['Disciplina'].isin(disciplinas_selecionadas)]
    
    if 'Data de Alteração' in tabela.columns:
        tabela['Data de Alteração'] = pd.to_datetime(tabela['Data de Alteração'], errors='coerce').dt.date
        
        data_inicial = data_range[0]
        data_final = data_range[1]

        tabela = tabela[(tabela['Data de Alteração'] >= data_inicial) & (tabela['Data de Alteração'] <= data_final)]
        
        tabela = tabela.drop(columns=['Data de Alteração'])

    return tabela

def ajusta_planilha_template(sigla_empreendimento, nome_empreendimento, arquivo_origem, arquivo_destino):
    novo_arquivo = SAIDAS_FOLDER / f"[{sigla_empreendimento}]-Lista_Mestra.xlsx"

    intervalo_inicial = "A2"
    intervalo_final = "M10000"

    wb_origem = load_workbook(arquivo_origem)
    ws_origem = wb_origem["Sheet1"]
    celulas_origem = ws_origem[intervalo_inicial:intervalo_final]

    wb_destino = load_workbook(arquivo_destino)
    ws_destino = wb_destino["0-DADOS"]

    linha_inicio = 2
    coluna_inicio = 1

    for i, linha in enumerate(celulas_origem):
        for j, celula in enumerate(linha):
            ws_destino.cell(row=linha_inicio + i, column=coluna_inicio + j, value=celula.value)

    ws_resumo = wb_destino["RESUMO GERAL"]
    ws_resumo['C4'] = nome_empreendimento

    wb_destino.save(novo_arquivo)
    return novo_arquivo

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/Assets/<path:filename>')
def custom_static_assets(filename):
    return send_file(os.path.join('Assets', filename))

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'Nenhum arquivo enviado'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Nenhum arquivo selecionado'}), 400
        
    if file and file.filename.endswith('.xlsx'):
        filename = f"{uuid.uuid4()}.xlsx"
        filepath = UPLOAD_FOLDER / filename
        file.save(filepath)
        
        try:
            tabela_tratada = pd.read_excel(filepath, skiprows=2)
            
            fases_unicas = []
            if "Fase" in tabela_tratada.columns:
                fases_unicas = tabela_tratada["Fase"].dropna().unique().tolist()
            
            disciplinas_unicas = []
            if "Disciplina" in tabela_tratada.columns:
                disciplinas_unicas = tabela_tratada["Disciplina"].dropna().unique().tolist()
            
            fases_unicas = [str(f) for f in fases_unicas if str(f) != "Fase"]
            disciplinas_unicas = [str(d) for d in disciplinas_unicas if str(d).strip() != "Disciplina" and str(d).strip() != ""]
            
            wb = openpyxl.load_workbook(filepath, read_only=True)
            sheet = wb.active
            nome_empreendimento = str(sheet['H4'].value) if sheet['H4'].value else "Empreendimento"
            
            # Extract the oldest date from 'Data de Alteração' column
            data_inicial_padrao = None
            col_data = 'Data de Alteração'
            if col_data in tabela_tratada.columns:
                datas = pd.to_datetime(tabela_tratada[col_data], dayfirst=True, errors='coerce').dropna()
                if not datas.empty:
                    data_inicial_padrao = datas.min().strftime('%Y-%m-%d')
            
            return jsonify({
                'success': True,
                'file_id': filename,
                'fases': sorted(fases_unicas),
                'disciplinas': sorted(disciplinas_unicas),
                'nome_empreendimento': nome_empreendimento,
                'data_inicial_padrao': data_inicial_padrao
            })
        except Exception as e:
            if filepath.exists():
                filepath.unlink()
            return jsonify({'error': str(e)}), 500
            
    return jsonify({'error': 'Arquivo inválido. Por favor, envie um arquivo .xlsx'}), 400

@app.route('/preview', methods=['POST'])
def preview_file():
    data = request.json
    file_id = data.get('file_id')
    fases_selecionadas = data.get('fases', [])
    disciplinas_selecionadas = data.get('disciplinas_selecionadas', [])
    data_inicial_str = data.get('data_inicial')
    data_final_str = data.get('data_final')
    
    if not file_id or not (UPLOAD_FOLDER / file_id).exists():
        return jsonify({'error': 'Arquivo não encontrado. Faça o upload novamente.'}), 400
        
    filepath = UPLOAD_FOLDER / file_id
    
    try:
        data_inicial = datetime.datetime.strptime(data_inicial_str, '%Y-%m-%d').date()
        data_final = datetime.datetime.strptime(data_final_str, '%Y-%m-%d').date()
        
        tabela_tratada = ajusta_lista(filepath, fases_selecionadas, disciplinas_selecionadas=disciplinas_selecionadas, data_range=[data_inicial, data_final])
        
        # Converter DataFrame para lista de dicionários (JSON) para envio ao frontend
        # Substitui NaNs por strings vazias para enviar um JSON válido
        preview_data = tabela_tratada.fillna('').to_dict(orient='records')
        
        return jsonify({
            'success': True,
            'data': preview_data
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/process', methods=['POST'])
def process_file():
    data = request.json
    file_id = data.get('file_id')
    fases_selecionadas = data.get('fases', [])
    disciplinas_selecionadas = data.get('disciplinas_selecionadas', [])
    data_inicial_str = data.get('data_inicial')
    data_final_str = data.get('data_final')
    tabela_editada = data.get('tabela_editada', []) # Recebe as edições das colunas I a M
    
    if not file_id or not (UPLOAD_FOLDER / file_id).exists():
        return jsonify({'error': 'Arquivo não encontrado. Faça o upload novamente.'}), 400
        
    filepath = UPLOAD_FOLDER / file_id
    
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True)
        sheet = wb.active
        nome_empreendimento = str(sheet['H4'].value) if sheet['H4'].value else "Empreendimento"
        sigla_empreendimento = nome_empreendimento[0:7] if len(nome_empreendimento) >= 7 else "PROJETO"

        data_inicial = datetime.datetime.strptime(data_inicial_str, '%Y-%m-%d').date()
        data_final = datetime.datetime.strptime(data_final_str, '%Y-%m-%d').date()
        
        tabela_tratada = ajusta_lista(filepath, fases_selecionadas, disciplinas_selecionadas=disciplinas_selecionadas, data_range=[data_inicial, data_final])
        
        # Injetar os dados editados pelo usuário no DataFrame
        if tabela_editada:
            # Definir as colunas na ordem I, J, K, L, M
            novas_colunas = [
                'Liberado para Orçamento', 
                'Liberado para compra de materiais?', 
                'Observações', 
                'Prazo de entrega dos projetos', 
                'Prazo de análise dos projetos'
            ]
            
            # Garantir que tabela_editada tenha o mesmo tamanho que tabela_tratada para criar o DataFrame alinhado
            # E caso algum id venha menor, preencher com dicts vazios
            while len(tabela_editada) < len(tabela_tratada):
                tabela_editada.append({c: '' for c in novas_colunas})
                
            edicoes_df = pd.DataFrame(tabela_editada[:len(tabela_tratada)])
            
            # Concatenar horizontalmente no DataFrame original
            tabela_tratada = pd.concat([tabela_tratada.reset_index(drop=True), edicoes_df.reset_index(drop=True)], axis=1)
        
        temp_tratada_path = UPLOAD_FOLDER / f"tratada_{file_id}"
        tabela_tratada.to_excel(temp_tratada_path, index=False)
        
        arquivo_final = ajusta_planilha_template(sigla_empreendimento, nome_empreendimento, temp_tratada_path, TEMPLATE_PATH)
        
        if temp_tratada_path.exists():
            temp_tratada_path.unlink()
            
        if filepath.exists():
            filepath.unlink()
            
        return jsonify({
            'success': True,
            'download_url': f'/download/{arquivo_final.name}'
        })
        
    except Exception as e:
        if filepath.exists():
            filepath.unlink()
        return jsonify({'error': str(e)}), 500

@app.route('/download/<filename>')
def download_file(filename):
    file_path = SAIDAS_FOLDER / filename
    if file_path.exists():
        @after_this_request
        def remove_file(response):
            try:
                if file_path.exists():
                    file_path.unlink()
            except Exception as e:
                app.logger.error("Erro ao deletar o arquivo: %s", e)
            return response

        return send_file(file_path, as_attachment=True)
    return jsonify({'error': 'Arquivo não encontrado'}), 404

if __name__ == '__main__':
    app.run(debug=True, port=8000)
