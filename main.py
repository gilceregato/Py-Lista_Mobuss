import streamlit as st
import pandas as pd
import tempfile
import openpyxl
import time
from pathlib import Path
from openpyxl import load_workbook

TEMPLATE_PATH = Path("Entradas/TEMPLATE-SITUAÇÃO DE PROJETOS COMPLEMENTARES.xlsx")

## Funções
def ajusta_lista(arquivo_excel, fases_selecionadas, remover_disciplina_arq=True):
    # Abrindo o arquivo Excel
    tabela = pd.read_excel(arquivo_excel)

    #Renomeando colunas
    tabela = tabela[['Revisão de Projeto',
                    'Unnamed: 1',
                    'Unnamed: 2',
                    'Unnamed: 3',
                    'Unnamed: 4',
                    'Unnamed: 5',
                    'Unnamed: 8',
                    'Unnamed: 9']]
    index_tabela = {'Revisão de Projeto':'Documento',
                    'Unnamed: 1':'Código',
                    'Unnamed: 2':'Revisão',
                    'Unnamed: 3':'Extensão',
                    'Unnamed: 4':'Situação',
                    'Unnamed: 5':'Título',
                    'Unnamed: 8':'Disciplina',
                    'Unnamed: 9':'Fase'}
    tabela.rename(columns=index_tabela, inplace=True)

    #Removendo linhas desnecessárias
    tabela = tabela.drop(tabela.index[0])
    tabela = tabela.drop(tabela.index[0])

    # Selecionando colunas
    tabela = tabela[['Disciplina',
                    'Fase',
                    'Código',
                    'Revisão',
                    'Documento',
                    'Extensão',
                    'Situação',
                    'Título']]

    # Filtrando por fase
    tabela = tabela[tabela['Fase'].isin(fases_selecionadas)]

    # Removendo disciplinas de ARQ, LEG e Incorporação
    disciplina_remover = ['ACE - Acessibilidade', 'ARQ - Arquitetônico', 'COM - Comunicação visual','INT - Interiores','PAI - Paisagismo','PUB - PUBLICIDADE']
    if remover_disciplina_arq == True:
        tabela = tabela[~tabela['Disciplina'].isin(disciplina_remover)]
    
    return tabela

def ajusta_planilha_template (sigla_empreendimento, nome_empreendimento,arquivo_origem, arquivo_destino):
    
    # Caminhos dos arquivos
    novo_arquivo = Path(rf"Saidas/[{sigla_empreendimento}]-Lista_Mestra.xlsx")

    # Intervalo de células que será copiado
    intervalo_inicial = "A2"
    intervalo_final = "H10000"

    # Abrir planilha de origem
    wb_origem = load_workbook(arquivo_origem)
    ws_origem = wb_origem["Sheet1"]

    # Obter linhas e colunas do intervalo
    celulas_origem = ws_origem[intervalo_inicial:intervalo_final]

    # Abrir planilha de destino
    wb_destino = load_workbook(arquivo_destino)
    ws_destino = wb_destino["0-DADOS"]

    # Ponto inicial para colar (A2 → linha 2, coluna 1)
    linha_inicio = 2
    coluna_inicio = 1

    # Copiar valores para o destino
    for i, linha in enumerate(celulas_origem):
        for j, celula in enumerate(linha):
            ws_destino.cell(row=linha_inicio + i, column=coluna_inicio + j, value=celula.value)

    # Substituir o nome do empreendimento na célula C4 da planilha destino, aba "RESUMO GERAL"
    ws_resumo = wb_destino["RESUMO GERAL"]
    ws_resumo['C4'] = nome_empreendimento

    # --- Salvar novo arquivo ---
    wb_destino.save(novo_arquivo)

    return novo_arquivo

########
if __name__ == '__main__':
    # Configurando a página inicial
    st.set_page_config(
        layout= 'wide', #formato da página,
        page_title='Gerar lista mestra de arquivos' #título da página
    )

    # Cabeçalho
    st.image('Assets/Logo_rottas.png',width=200)
    st.title('Gerar lista mestra de arquivos')

    st.write('Esta ferramenta permite gerar uma lista mestra de arquivos de projetos complementares a partir de um arquivo Excel fornecido pelo usuário. Basta fazer o upload do arquivo, selecionar as fases desejadas e clicar em "Gerar lista mestra".')
    st.divider()

    
    # Cria botão para upload de arquivo excel
    st.markdown("### 1) Faça upload do arquivo do Mobuss:")
    uploaded_file = st.file_uploader("Escolha o arquivo Excel", type=["xlsx"])

    if uploaded_file:
        with st.spinner(f"Processando: {uploaded_file.name}", show_time=True):
            time.sleep(5)

        # 1. Criar arquivo temporário para salvar a tabela tratada
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
            temp_tratada_path = Path(temp_file.name)
        
        # 2. Ler o arquivo excel enviado e listar as fases de projeto dele
        tabela_tratada = pd.read_excel(uploaded_file)
        if "Fase" in tabela_tratada.columns:
            fases_unicas = tabela_tratada["Fase"].dropna().unique()
        elif "Unnamed: 9" in tabela_tratada.columns:
            fases_unicas = tabela_tratada["Unnamed: 9"].dropna().unique()
        else:
            fases_unicas = []
        
        # Remove "Fase" da lista, se existir
        fases_unicas = [f for f in fases_unicas if f != "Fase"]

        # Coleta o nome do projeto/empreendimento
        wb =openpyxl.load_workbook(uploaded_file, read_only=True)
        sheet = wb.active
        nome_empreendimento = sheet['H4'].value
        sigla_empreendimento = nome_empreendimento[0:7]

        # Seleção de fases
        st.markdown("### 2) Seleção de Fases")
        fases_selecionadas = st.multiselect(
            label="**Selecione as fases que deseja incluir na lista mestra gerada a partir do arquivo enviado:**",
            options=sorted(fases_unicas),
            default=list(fases_unicas)
        )
        if not fases_selecionadas:
            st.warning("Selecione pelo menos uma fase para continuar.")
        else:

            # Botão para iniciar o processamento
            st.divider()
            st.markdown("### 3) Gerar Lista Mestra")

            st.markdown("3.1 - Se a caixa abaixo estiver marcada as disciplinas de Acessibilidade, Arquitetura, Comunicação Visual, Interiores, Paisagismo e Publicidade serão removidas da lista mestra gerada.")
            remover_disciplina_arq = st.checkbox("Remover disciplinas de Arquitetura", value=True)
            st.markdown("")
            st.markdown("3.2 - Após clicar no botão abaixo, aguarde o processamento e, em seguida, faça o download da planilha ajustada.")
            gerar = st.button("Gerar lista mestra")
            if gerar:
                # 2. Ajustar a lista e salvar arquivo tratado temporário
                tabela_tratada = pd.read_excel(uploaded_file)
                tabela_tratada = ajusta_lista(uploaded_file, fases_selecionadas, remover_disciplina_arq=True)
                tabela_tratada.to_excel(temp_tratada_path, index=False)

                # 3. Aplicar o template na planilha tratada
                arquivo_final = ajusta_planilha_template(sigla_empreendimento,nome_empreendimento,temp_tratada_path, TEMPLATE_PATH)

                # 4. Excluir o arquivo temporário
                if temp_tratada_path.exists():
                    temp_tratada_path.unlink()
                
                # 5. Botão para download do arquivo final
                st.divider()
                st.markdown("### 4) Baixar arquivo da Lista Mestra")
                with open(arquivo_final, "rb") as f:
                    st.download_button(
                        label="Baixar Planilha Ajustada",
                        data=f,
                        file_name=arquivo_final.name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                
                st.success("Processamento concluído com sucesso! Faça o download da planilha ajustada.")
        
        
    # Rodapé
    st.divider()
    st.markdown(':red[Atenção: os documentos gerados por este script, ou parte deles, é de **uso exclusivo para uso da Rottas Construtora e Incorporadora**. Seu envio para terceiros, bem como sua reprodução total ou parcial são proibidos.]')
    st.write('Criado por Gilmar Ceregato | 2025')
    st.write('Rev. 01 - 22/10/2025')
